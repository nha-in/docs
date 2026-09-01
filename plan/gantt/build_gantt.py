#!/usr/bin/env python3
"""Build the ABDM Developer Portal V1 gantt workbook.

The task table below is compiled from plan#p8-schedule. It is not a second
source. Change the plan first, then this file, then restamp PLAN_VERSION.

Live status (% Done, Status) is NOT held here. It lives in the Google Sheet,
which is the ledger. Pass --status with a CSV export of the Gantt tab to carry
it forward into a rebuild, so regenerating never wipes progress.

    python3 gantt/build_gantt.py                       # fresh, everything at 0%
    python3 gantt/build_gantt.py --status live.csv     # keep current progress
    python3 gantt/build_gantt.py --check               # self-check only

Import the result into the existing Sheet with File > Import > Replace
spreadsheet. That keeps the URL and the sharing, which a re-upload would not.
"""

import argparse
import csv
import sys
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as L
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

PLAN_VERSION = "2026.08.24-4"
SHEET_URL = "https://docs.google.com/spreadsheets/d/17OkDm9-wUSgMZ5G2Tegv69-MP-aXTdjjjN37t4-HxSE/edit"

NAVY, BAND, PLAN = "14274E", "E8EDF7", "A9C3F0"
DONE, MILE, WKND = "34A853", "6B3FA0", "F1F3F4"
TODAY_C, GREY, LINE = "FDE7E9", "5F6368", "D0D7E2"

STATUSES = "Not started,In progress,Blocked,Done"
STATUS_COLOURS = [
    ("Done", "E6F4EA", "137333"),
    ("In progress", "E8F0FE", "1967D2"),
    ("Blocked", "FCE8E6", "C5221F"),
    ("Not started", "F1F3F4", "5F6368"),
]

# wbs, task, stream, owner, start, days. days=0 is a milestone.
# A stream row (wbs with no dot) rolls up from its children; its dates are formulas.
TASKS = [
    ("1",   "Catalogue",                 "Catalogue", "Both",      None,            None),
    ("1.1", "Atom schema and lint",      "Catalogue", "Product",   date(2026, 8, 24), 2),
    ("1.2", "Ingest HIE-CM specs",       "Catalogue", "Shyamjith", date(2026, 8, 25), 2),
    ("1.3", "Write M1 atoms",            "Catalogue", "Both",      date(2026, 8, 26), 2),
    ("1.4", "Write M2 M3 atoms",         "Catalogue", "Both",      date(2026, 8, 28), 3),
    ("1.5", "Write shared atoms",        "Catalogue", "Both",      date(2026, 8, 31), 2),
    ("1.6", "Verify and fix gaps",       "Catalogue", "Both",      date(2026, 9, 1),  2),
    ("2",   "Site and MCP",              "Site and MCP", "Shyamjith", None,            None),
    ("2.1", "Set up self hosted site",   "Site and MCP", "Shyamjith", date(2026, 8, 25), 2),
    ("2.2", "Module specs and nav",      "Site and MCP", "Shyamjith", date(2026, 8, 27), 1),
    ("2.3", "Build Docs MCP server",     "Site and MCP", "Shyamjith", date(2026, 8, 28), 1),
    ("2.4", "Wire deploys and publish",  "Site and MCP", "Shyamjith", date(2026, 8, 29), 1),
    ("3",   "Skills",                    "Skills",    "Both",      None,            None),
    ("3.1", "Build skill compiler",      "Skills",    "Shyamjith", date(2026, 8, 27), 3),
    ("3.2", "Build index and plugin",    "Skills",    "Shyamjith", date(2026, 8, 30), 1),
    ("3.3", "Compile M1 skills",         "Skills",    "Both",      date(2026, 8, 31), 2),
    ("3.4", "Compile all skills",        "Skills",    "Shyamjith", date(2026, 9, 2),  2),
    ("4",   "Pipeline",                  "Pipeline",  "Shyamjith", None,            None),
    ("4.1", "Build source watcher",      "Pipeline",  "Shyamjith", date(2026, 8, 29), 2),
    ("4.2", "Set up CI checks",          "Pipeline",  "Shyamjith", date(2026, 8, 31), 2),
    ("4.3", "Publish llms.txt file",     "Pipeline",  "Shyamjith", date(2026, 9, 2),  1),
    ("5",   "Proof",                     "Proof",     "Both",      None,            None),
    ("5.1", "Wire support agent",        "Proof",     "Shyamjith", date(2026, 9, 1),  2),
    ("5.2", "Run first day test",        "Proof",     "Product",   date(2026, 9, 3),  1),
    ("5.3", "Run evals and fix",         "Proof",     "Both",      date(2026, 9, 3),  2),
    ("6",   "Ship V1",                   "Ship",      "Both",      date(2026, 9, 5),  0),
]

CHECKPOINTS = [
    (date(2026, 8, 26), "Open the docs site with the NHA HIE-CM OpenAPI references rendered and searchable, plus the sandbox registration guide.", "Shyamjith"),
    (date(2026, 8, 28), "Read dummy proof M1 pages, run the M1 curls against sandbox, and ask the Docs MCP questions.", "Both"),
    (date(2026, 8, 30), "Install the abdm-index and hiecm-m1 skills into Claude Code and scaffold M1.", "Shyamjith"),
    (date(2026, 9, 2),  "Read M2 and M3 at the same depth, install the full plugin, and see the phase scope stated on every landing page.", "Both"),
    (date(2026, 9, 5),  "Use the whole thing, file a gap from the support agent, and see an NHA source change open a pull request.", "Both"),
]

DONE_CRITERIA = [
    (1,  "Catalogue lint passes on main: schema valid, five sections present, no em dash, all related ids resolve, all sources hashed.", "Product"),
    (2,  "Every HIE-CM M1 to M3 endpoint atom has a curl run against sandbox on or after 28 August, with the response recorded.", "Shyamjith"),
    (3,  "Every NHA functional test case for M1 to M3 exists as a test atom and is referenced by a test skill.", "Product"),
    (4,  "All skills compile, validate and install individually into Claude Code. The plugin installs as one unit.", "Shyamjith"),
    (5,  "The abdm-index skill is generated from the graph and lists every skill, agent and tool.", "Shyamjith"),
    (6,  "Docs site live with search and the module references. Docs MCP deployed with nine tools over the current snapshot.", "Shyamjith"),
    (7,  "The watcher has opened at least one real pull request from a real NHA source change.", "Shyamjith"),
    (8,  "The support agent answered the six eval tasks from the Catalogue, citing atom ids, with the score recorded.", "Both"),
    (9,  "First day developer test: no ABDM exposure, docs plus sandbox credentials only, a working M1 call in under two hours with no human help.", "Product"),
    (10, "Landing page, index entries and skill descriptions state the phase scope: M1 to M3 in Phase 1, M4 and UHI in Phase 2, NHCX out. Lint rejects gateway: nhcx.", "Product"),
    (11, "Public repo, neutral licence, CONTRIBUTING, SECURITY and GOVERNANCE files, and no eka.care reference in the core Catalogue.", "Product"),
]

ACTIONS = [
    (1, date(2026, 8, 24), "Apply for sandbox credentials today. They take three to four days.", "Shyamjith", date(2026, 8, 24)),
    (2, date(2026, 8, 24), "Sign off the phasing call: HIE-CM M1 to M3 in Phase 1, M4 and UHI in Phase 2, NHCX out of scope.", "Product", date(2026, 8, 25)),
    (3, date(2026, 8, 24), "Reach out to OHCN about abdm-docs.pages.dev. Propose the Catalogue as shared upstream.", "Product", date(2026, 8, 26)),
    (4, date(2026, 8, 24), "Accept that some NHA endpoints ship unverified in V1 and are labelled as such.", "Product", date(2026, 8, 26)),
    (5, date(2026, 8, 24), "Decided: fully self hosted from day one. Docusaurus with Scalar MIT packages, own Docs MCP, Ollama sidecar.", "Product", date(2026, 8, 26)),
    (6, date(2026, 8, 24), "Decide the licence and the public repo home before the first push.", "Product", date(2026, 8, 25)),
    (7, date(2026, 8, 24), "Recruit the first day developer for the 3 September test.", "Product", date(2026, 8, 31)),
    (8, date(2026, 8, 24), "Name a backup for Shyamjith. He owns twelve of twenty one tasks alone and is on nineteen.", "Product", date(2026, 8, 26)),
]

TAIL_BUFFER_DAYS = 3        # blank days after the last due date, so slip has somewhere to show
FIRST_DAY_COL = 10          # column J
HDR = 8                     # table header row
D0 = 9                      # first data row

thin = Side(style="thin", color=LINE)
box = Border(left=thin, right=thin, top=thin, bottom=thin)


def fill(colour):
    return PatternFill("solid", fgColor=colour)


def day_span():
    """First and last day column, derived from the tasks so a slip widens the grid."""
    starts = [t[4] for t in TASKS if t[4]]
    ends = [t[4] + timedelta(days=max(t[5] - 1, 0)) for t in TASKS if t[4]]
    start = min(starts)
    return start, (max(ends) - start).days + 1 + TAIL_BUFFER_DAYS


def read_status(path):
    """Pull {wbs: (percent, status)} out of a CSV export of the live Gantt tab."""
    carried = {}
    with open(path, newline="") as fh:
        for row in csv.reader(fh):
            if len(row) < 9 or "." not in row[0]:
                continue
            pct = row[7].strip().rstrip("%")
            try:
                carried[row[0].strip()] = (float(pct) / 100, row[8].strip() or "Not started")
            except ValueError:
                continue
    if not carried:
        sys.exit(f"no task rows found in {path}. Export the Gantt tab as CSV, not the whole book.")
    return carried


def build_gantt(ws, carried):
    ws.sheet_view.showGridLines = False
    start, days = day_span()
    last = D0 + len(TASKS) - 1
    J = L(FIRST_DAY_COL)

    for col, w in [("A", 7), ("B", 30), ("C", 12), ("D", 12), ("E", 11), ("F", 7), ("G", 11), ("H", 9), ("I", 13)]:
        ws.column_dimensions[col].width = w
    for i in range(days):
        ws.column_dimensions[L(FIRST_DAY_COL + i)].width = 4.2

    ws["A2"] = "ABDM Developer Portal, V1 Phase 1"
    ws["A2"].font = Font(bold=True, size=16, color=NAVY)
    ws.merge_cells("A2:D2")
    ws["A3"] = "One catalogue of HIE-CM, M1 to M3, at full depth. Docs, MCP and skills built from it."
    ws["A3"].font = Font(size=10, color=GREY)
    ws.merge_cells("A3:D3")

    for ref, val in [("A5", "OWNERS"), ("B5", "Product and Shyamjith"), ("A6", "START"), ("B6", start),
                     ("C5", "SHIP"), ("D5", date(2026, 9, 5)), ("C6", "TODAY"), ("D6", "=TODAY()")]:
        ws[ref] = val
        ws[ref].font = (Font(bold=True, size=9, color=GREY) if ref[0] in "AC"
                        else Font(bold=True, size=10, color=NAVY))
    for ref in ("B6", "D5", "D6"):
        ws[ref].number_format = "dd mmm yyyy"

    for ref, label, formula, fmt in [
        ("E5", "% COMPLETE", f"=IFERROR(SUMPRODUCT($F${D0}:$F${last},$H${D0}:$H${last})/SUM($F${D0}:$F${last}),0)", "0%"),
        ("E6", "DAYS TO SHIP", "=$D$5-TODAY()", "0"),
        ("G5", "TASKS", f'=COUNTIF($A${D0}:$A${last},"?.*")', "0"),
        ("G6", "BLOCKED", f'=COUNTIF($I${D0}:$I${last},"Blocked")', "0"),
    ]:
        ws[ref] = label
        ws[ref].font = Font(bold=True, size=9, color=GREY)
        cell = ws.cell(row=int(ref[1]), column=ws[ref].column + 1, value=formula)
        cell.number_format = fmt
        cell.font = Font(bold=True, size=12, color=NAVY)

    ws.cell(row=2, column=FIRST_DAY_COL, value="LEGEND").font = Font(bold=True, size=9, color=GREY)
    for i, (label, colour) in enumerate([("Planned", PLAN), ("Done", DONE), ("Weekend", WKND), ("Milestone", MILE)]):
        r = 3 + i
        sw = ws.cell(row=r, column=FIRST_DAY_COL)
        sw.fill, sw.border = fill(colour), box
        ws.cell(row=r, column=FIRST_DAY_COL + 1, value=label).font = Font(size=9, color=GREY)
        ws.merge_cells(start_row=r, start_column=FIRST_DAY_COL + 1, end_row=r, end_column=FIRST_DAY_COL + 4)

    for i in range(days):
        c = FIRST_DAY_COL + i
        weekday = ws.cell(row=7, column=c, value=f'=TEXT({L(c)}8,"ddd")')
        daynum = ws.cell(row=8, column=c, value=start + timedelta(days=i))
        daynum.number_format = "d"
        for cell in (weekday, daynum):
            cell.fill = fill(NAVY)
            cell.font = Font(bold=True, size=9, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
            cell.border = box

    label = f"{start.strftime('%d %b').lstrip('0')} to {(start + timedelta(days=days - 1)).strftime('%d %b %Y').lstrip('0')}"
    ws.cell(row=7, column=2, value=label).font = Font(bold=True, size=9, color=GREY)
    for i, head in enumerate(["WBS", "Task", "Stream", "Owner", "Start", "Days", "Due", "% Done", "Status"], start=1):
        c = ws.cell(row=HDR, column=i, value=head)
        c.fill = fill(NAVY)
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.alignment = Alignment(horizontal="center" if i > 3 else "left", vertical="center")
        c.border = box
    ws.row_dimensions[HDR].height = 22

    children = {}
    for idx, (wbs, _, stream, _, _, _) in enumerate(TASKS):
        if "." not in wbs and wbs != TASKS[-1][0]:
            children[D0 + idx] = [D0 + j for j, t in enumerate(TASKS) if t[2] == stream and "." in t[0]]

    for idx, (wbs, name, stream, owner, s, d) in enumerate(TASKS):
        r = D0 + idx
        is_stream = r in children
        pct, status = carried.get(wbs, (0, "Not started"))
        ws.cell(row=r, column=1, value=wbs)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=stream)
        ws.cell(row=r, column=4, value=owner)
        if is_stream:
            lo, hi = min(children[r]), max(children[r])
            ws.cell(row=r, column=5, value=f"=MIN(E{lo}:E{hi})")
            ws.cell(row=r, column=7, value=f"=MAX(G{lo}:G{hi})")
            ws.cell(row=r, column=6, value=f"=G{r}-E{r}+1")
            ws.cell(row=r, column=8, value=f"=IFERROR(SUMPRODUCT(F{lo}:F{hi},H{lo}:H{hi})/SUM(F{lo}:F{hi}),0)")
        else:
            ws.cell(row=r, column=5, value=s)
            ws.cell(row=r, column=6, value=d)
            ws.cell(row=r, column=7, value=f"=IF(F{r}=0,E{r},E{r}+F{r}-1)")
            ws.cell(row=r, column=8, value=pct)
            ws.cell(row=r, column=9, value=status)

        for col in range(1, FIRST_DAY_COL + days):
            c = ws.cell(row=r, column=col)
            c.border = box
            if is_stream:
                c.fill = fill(BAND)
                c.font = Font(bold=True, size=10, color=NAVY)
            elif d == 0:
                c.font = Font(bold=True, size=10, color=MILE)
            else:
                c.font = Font(size=10)
            if col in (1, 5, 6, 7, 8, 9) or col >= FIRST_DAY_COL:
                c.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=5).number_format = "dd mmm"
        ws.cell(row=r, column=7).number_format = "dd mmm"
        ws.cell(row=r, column=8).number_format = "0%"
        ws.row_dimensions[r].height = 19

    grid = f"{J}{D0}:{L(FIRST_DAY_COL + days - 1)}{last}"
    for formula, colour in [
        (f"AND($F{D0}=0,{J}$8=$E{D0})", MILE),
        (f"AND($F{D0}>0,$H{D0}>0,{J}$8>=$E{D0},{J}$8<=$E{D0}+ROUND($F{D0}*$H{D0},0)-1)", DONE),
        (f"AND($F{D0}>0,{J}$8>=$E{D0},{J}$8<=$G{D0})", PLAN),
        (f"{J}$8=TODAY()", TODAY_C),
        (f"WEEKDAY({J}$8,2)>5", WKND),
    ]:
        ws.conditional_formatting.add(grid, FormulaRule(formula=[formula], fill=fill(colour), stopIfTrue=True))

    head_range = f"{J}7:{L(FIRST_DAY_COL + days - 1)}8"
    ws.conditional_formatting.add(head_range, FormulaRule(
        formula=[f"{J}$8=TODAY()"], fill=fill("C5221F"), font=Font(bold=True, color="FFFFFF"), stopIfTrue=True))
    ws.conditional_formatting.add(head_range, FormulaRule(
        formula=[f"WEEKDAY({J}$8,2)>5"], fill=fill("7A879B"), stopIfTrue=True))

    status_range = f"I{D0}:I{last}"
    for value, bg, fg in STATUS_COLOURS:
        ws.conditional_formatting.add(status_range, FormulaRule(
            formula=[f'$I{D0}="{value}"'], fill=fill(bg), font=Font(bold=True, color=fg), stopIfTrue=True))
    ws.conditional_formatting.add(f"B{D0}:B{last}", FormulaRule(
        formula=[f'AND($G{D0}<TODAY(),$H{D0}<1,$A{D0}<>"",ISNUMBER($G{D0}))'],
        font=Font(bold=True, color="C5221F"), stopIfTrue=True))
    ws.conditional_formatting.add(f"H{D0}:H{last}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FFFFFF",
        end_type="num", end_value=1, end_color="B7E1C4"))

    dv = DataValidation(type="list", formula1=f'"{STATUSES}"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(status_range)

    ws.auto_filter.ref = f"A{HDR}:I{last}"
    ws.freeze_panes = f"{J}{D0}"
    return last


def add_table_sheet(wb, name, title, subtitle, headers, rows, widths, table_name,
                    date_cols=(), status_col=None):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=15, color=NAVY)
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=10, color=GREY)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[L(i)].width = w

    hr = 4
    for i, head in enumerate(headers, start=1):
        c = ws.cell(row=hr, column=i, value=head)
        c.fill = fill(NAVY)
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.alignment = Alignment(vertical="center")
        c.border = box
    ws.row_dimensions[hr].height = 22

    for ri, row in enumerate(rows, start=hr + 1):
        for ci, value in enumerate(row, start=1):
            c = ws.cell(row=ri, column=ci, value=value)
            c.border = box
            c.font = Font(size=10)
            c.alignment = Alignment(vertical="top", wrap_text=ci in (2, 3, 4))
            if ci in date_cols:
                c.number_format = "dd mmm yyyy"
                c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[ri].height = 30

    end = hr + len(rows)
    table = Table(displayName=table_name, ref=f"A{hr}:{L(len(headers))}{end}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
    ws.add_table(table)

    if status_col:
        rng = f"{status_col}{hr + 1}:{status_col}{end}"
        dv = DataValidation(type="list", formula1=f'"{STATUSES}"', allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(rng)
        for value, bg, fg in STATUS_COLOURS:
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'${status_col}{hr + 1}="{value}"'], fill=fill(bg),
                font=Font(bold=True, color=fg), stopIfTrue=True))
    ws.freeze_panes = f"A{hr + 1}"


def build(out, carried):
    wb = Workbook()
    last = build_gantt(wb.active, carried)
    wb.active.title = "Gantt"

    add_table_sheet(wb, "Checkpoints", "Checkpoints",
                    "What an integrator can actually do at the end of each increment.",
                    ["By end of", "An integrator can", "Owner", "Status"],
                    [(d, text, owner, "Not started") for d, text, owner in CHECKPOINTS],
                    [16, 78, 14, 14], "Checkpoints", date_cols=(1,), status_col="D")

    add_table_sheet(wb, "Definition of done", "Definition of done",
                    "Eleven checks. Every one is observable. None is a judgement call.",
                    ["#", "Check", "Owner", "Status"],
                    [(n, text, owner, "Not started") for n, text, owner in DONE_CRITERIA],
                    [6, 82, 14, 14], "DefinitionOfDone", status_col="D")

    add_table_sheet(wb, "Actions", "Actions and decisions",
                    "Open items from the risk table, plus anything raised in standup. Add rows as they come.",
                    ["No", "Raised", "Item", "Owner", "Due", "Status", "Notes"],
                    [(n, raised, text, owner, due, "Not started", "") for n, raised, text, owner, due in ACTIONS],
                    [6, 13, 60, 13, 13, 14, 30], "Actions", date_cols=(2, 5), status_col="F")

    wb.save(out)
    return last


def check():
    """One runnable check: the task table is sane and the grid covers every bar."""
    seen = set()
    for wbs, name, _, _, s, d in TASKS:
        assert wbs not in seen, f"duplicate WBS {wbs}"
        seen.add(wbs)
        assert len(name.split()) < 5, f"task name over four words: {name}"
        assert "—" not in name and "–" not in name, f"dash in task name: {name}"
        assert (s is None) == (d is None), f"{wbs} has a start without days, or the reverse"
    start, days = day_span()
    grid_end = start + timedelta(days=days - 1)
    for wbs, _, _, _, s, d in TASKS:
        if s:
            assert s >= start, f"{wbs} starts before the grid"
            assert s + timedelta(days=max(d - 1, 0)) <= grid_end, f"{wbs} ends past the grid"
    tasks = [t for t in TASKS if t[4] is not None]
    solo = sum(1 for t in tasks if t[3] == "Shyamjith")
    on = sum(1 for t in tasks if t[3] in ("Shyamjith", "Both"))
    assert (len(tasks), solo, on) == (21, 12, 19), (
        f"Actions row 8 claims 12 of 21 alone and 19 involved, table says "
        f"{solo} of {len(tasks)} and {on}. Update the wording or the owners.")
    for stream in {t[2] for t in TASKS if "." in t[0]}:
        assert any(t[2] == stream and "." not in t[0] for t in TASKS), f"{stream} has no rollup row"
    print(f"ok   {len(TASKS)} rows, {days} day columns, plan_version {PLAN_VERSION}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--status", metavar="CSV", help="CSV export of the live Gantt tab, to carry progress forward")
    ap.add_argument("--out", default="gantt/abdm-portal-v1-gantt.xlsx")
    ap.add_argument("--check", action="store_true", help="run the self-check and stop")
    args = ap.parse_args()

    check()
    if args.check:
        sys.exit(0)
    build(args.out, read_status(args.status) if args.status else {})
    print(f"ok   wrote {args.out}")
    print(f"     import into {SHEET_URL}")
    print("     File > Import > Replace spreadsheet, so the URL and sharing survive")
